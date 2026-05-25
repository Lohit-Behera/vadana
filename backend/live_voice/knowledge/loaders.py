"""Load document text via Docling (PDF/DOCX/XLSX) with legacy fallbacks."""

from __future__ import annotations

import logging
from collections.abc import Callable
from pathlib import Path
from typing import TYPE_CHECKING

from live_voice.hf_env import configure_hf_hub

configure_hf_hub()

logger = logging.getLogger(__name__)

SUPPORTED_SUFFIXES = {".md", ".pdf", ".docx", ".xlsx"}

if TYPE_CHECKING:
    from docling.document_converter import DocumentConverter

_converter: DocumentConverter | None = None


def _get_docling_converter(on_progress: Callable[[str], None] | None = None):
    global _converter
    if _converter is not None:
        return _converter
    if on_progress:
        on_progress("Loading Docling parser (first run may download layout models)…")
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption

    pdf_options = PdfPipelineOptions()
    pdf_options.do_ocr = False

    _converter = DocumentConverter(
        format_options={
            InputFormat.PDF: PdfFormatOption(pipeline_options=pdf_options),
        },
    )
    if on_progress:
        on_progress("Docling ready")
    return _converter


def _load_with_docling(path: Path, on_progress: Callable[[str], None] | None = None) -> str:
    converter = _get_docling_converter(on_progress)
    if on_progress:
        on_progress(f"Parsing {path.name} with Docling…")
    result = converter.convert(str(path))
    text = result.document.export_to_markdown() or ""
    return text.strip()


def _load_legacy(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        parts = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(parts).strip()
    if suffix == ".docx":
        import docx2txt

        return (docx2txt.process(str(path)) or "").strip()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            lines.append(f"## Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines).strip()
    return ""


def load_file_text(
    path: Path,
    *,
    on_progress: Callable[[str], None] | None = None,
    prefer_fast: bool = False,
) -> str:
    """Load document text. prefer_fast skips Docling (use for live voice turns)."""
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported file type: {suffix}")
    if suffix == ".md":
        return path.read_text(encoding="utf-8", errors="replace").strip()

    if prefer_fast:
        text = _load_legacy(path)
        if text.strip():
            return text
        logger.info("Fast parser empty for %s; trying Docling", path.name)

    try:
        text = _load_with_docling(path, on_progress)
        if text.strip():
            return text
        logger.warning("Docling returned no text for %s; using legacy parser", path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Docling failed for %s (%s); using legacy parser", path, exc)
        if on_progress:
            on_progress(f"Docling failed, using fallback parser for {path.name}…")
    return _load_legacy(path)


def reset_docling_converter() -> None:
    """Drop cached converter (e.g. after rebuild completes)."""
    global _converter
    _converter = None
